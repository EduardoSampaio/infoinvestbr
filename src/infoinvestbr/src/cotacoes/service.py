import requests
import yfinance as yf
from datetime import datetime
from sqlalchemy.orm import Session
import math
from src.cotacoes.schemas import CotacaoSchema, CotacaoAltaOuBaixaSchema
from src.core.schemas import DividendoSchema
from src.core import models
from src.core.models import HistoricoCotacao
from src.core.exceptions import CodigoAtivoException
from src.cotacoes.constantes import ACAO_TICKER, FUNDOS_TICKER
from openpyxl import load_workbook


def get_by_codigo(codigo: str, periodo: str, intervalo: str):
    """
     codigo  = [CODIGO.SA]
     periodo = [1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max]\n
     intevalo= [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
    """
    cotacao = yf.Ticker(f'{codigo}.SA')
    cotacao = cotacao.history(period=periodo, interval=intervalo)

    list_cotacao = []
    for data in cotacao['Close'].keys():
        _cotacao = CotacaoSchema(
            codigo=codigo,
            data=data,
            abertura=cotacao['Open'].get(data),
            fechamento=cotacao['Close'].get(data),
            alta=cotacao['High'].get(data),
            baixa=cotacao['Low'].get(data),
        )
        list_cotacao.append(_cotacao)

    return list_cotacao


def get_by_codigo_chart(codigo: str, periodo: str, intervalo: str):
    """
     codigo  = [CODIGO.SA]
     periodo = [1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max]\n
     intevalo= [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
    """
    cotacao = yf.Ticker(f'{codigo}.SA')
    cotacao = cotacao.history(period=periodo, interval=intervalo)

    list_data = []
    list_close = []
    for data in cotacao['Close'].keys():
        list_close.append(cotacao['Close'].get(data))
        list_data.append("{:%d/%m/%Y}".format(data))

    return {
        "datas": list_data,
        "fechamento": list_close
    }


def get_by_codigo_ibovespa(periodo: str, intervalo: str):
    """
     [IFiX.SA]
     periodo = [1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max]\n
     intevalo= [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
    """
    cotacao = yf.Ticker("^BVSP")
    cotacao = cotacao.history(period=periodo, interval=intervalo)

    list_cotacao = []
    for data in cotacao['Close'].keys():
        _cotacao = CotacaoSchema(
            codigo="^BVSP",
            data=data,
            abertura=cotacao['Open'].get(data),
            fechamento=cotacao['Close'].get(data),
            alta=cotacao['High'].get(data),
            baixa=cotacao['Low'].get(data),
        )
        list_cotacao.append(_cotacao)

    return list_cotacao


def get_by_codigo_by_intervalo(codigo, inicio: datetime.date, fim: datetime.date):
    """
     codigo  = [CODIGO.SA]
     periodo = [1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max]\n
     intevalo= [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
    """
    cotacao = yf.Ticker(f'{codigo}.SA')
    start = inicio.strftime('%Y-%m-%d')
    end = fim.strftime('%Y-%m-%d')
    cotacao = cotacao.history(start=start, end=end)
    list_cotacao = []
    for data in cotacao['Close'].keys():
        _cotacao = CotacaoSchema(
            codigo=codigo,
            data=data,
            abertura=cotacao['Open'].get(data),
            fechamento=cotacao['Close'].get(data),
            alta=cotacao['High'].get(data),
            baixa=cotacao['Low'].get(data),
        )
        list_cotacao.append(_cotacao)

    return list_cotacao


def get_by_codigo_atual(codigo: str) -> float:
    cotacao = yf.Ticker(f'{codigo}.SA')
    cotacao = cotacao.history(period="1d")['Close'][0]
    return cotacao


def get_by_codigo_varicao_diaria(codigo: str):
    cotacao = yf.Ticker(f'{codigo}.SA')
    fechamento_anterior = cotacao.history(period="2d")['Close'][-2]
    fechamento_atual = cotacao.history(period="1d")['Close'][0]
    return fechamento_atual, fechamento_anterior


def get_moeda_cotacao():
    request_url = requests.get("https://economia.awesomeapi.com.br/json/last/USD-BRL,EUR-BRL,BTC-BRL,ETH-BRL")
    request_url_dic = request_url.json()
    return {
        "dolar": request_url_dic["USDBRL"]["bid"],
        "euro": request_url_dic["EURBRL"]["bid"],
        "btc": request_url_dic["BTCBRL"]["bid"],
        "ethereum": request_url_dic["ETHBRL"]["bid"],
    }


def gerar_dados_historicos(db: Session, codigo: str, periodo: str = "1d", updated: bool = False):
    """
         codigo  = [CODIGO.SA]
         periodo = [1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max]\n
    """
    ativo = db.query(models.HistoricoCotacao).filter(models.HistoricoCotacao.codigo == codigo).first()

    if ativo is None:
        create_historico(db, codigo, periodo)
    if updated:
        deletar_dados_historicos_by_codigo(db, codigo)
        create_historico(db, codigo, periodo)


def deletar_dados_historicos(db: Session):
    """

    :param db:
    """
    valores = db.query(models.HistoricoCotacao).all()
    for valor in valores:
        db.delete(valor)

    db.commit()


def deletar_dados_historicos_by_codigo(db: Session, codigo: str):
    """

    :param db:
    :param codigo:
    """
    valores = db.query(models.HistoricoCotacao).filter(models.HistoricoCotacao.codigo == codigo).all()
    for valor in valores:
        db.delete(valor)

    db.commit()


def create_historico(db: Session, codigo: str, periodo: str = "1d"):
    """

    :param db:
    :param codigo:
    :param periodo:
    :return:
    """
    ticker = yf.Ticker(f'{codigo}.SA')
    if ticker is None:
        return None
    else:
        cotacoes = ticker.history(period=periodo)['Close']
        for data in cotacoes.keys():
            valor = cotacoes.get(data)
            historico = HistoricoCotacao(codigo=codigo, data=data, valor=valor, periodo=periodo)
            db.add(historico)
            db.commit()
            db.refresh(historico)


def get_historico_dividendo(codigo: str):
    current_ticker = yf.Ticker(f'{codigo}.SA')
    valores = current_ticker.dividends
    list_dividendos = []

    my_dict = {}
    for data in current_ticker.dividends.keys():
        # ativo = DividendoSchema(codigo, data, valores.get(data))
        # list_dividendos.append(ativo)
        ano = "{:%d/%m/%Y}".format(data).split('/')[2]
        my_dict.setdefault(ano, []).append(valores.get(data))

    list_data = []
    lista_valores = []
    for k in my_dict.keys():
        valores = my_dict.get(k)
        total = 0
        for valor in valores:
            total += valor
        list_data.append(k)
        lista_valores.append(round(total, 2))

    return {
        "datas": list_data,
        "valores": lista_valores
    }


def codigo_exception(codigo: str):
    if codigo.find(".SA") == -1:
        raise CodigoAtivoException(name=codigo)


def get_maiores_altas_dia_acoes() -> CotacaoAltaOuBaixaSchema:
    return get_cotacao_comparacao(True, ACAO_TICKER)


def get_maiores_altas_dia_fundos() -> CotacaoAltaOuBaixaSchema:
    return get_cotacao_comparacao(True, FUNDOS_TICKER)


def get_maiores_baixa_dia_acoes() -> CotacaoAltaOuBaixaSchema:
    return get_cotacao_comparacao(False, ACAO_TICKER)


def get_maiores_baixa_dia_fundos() -> CotacaoAltaOuBaixaSchema:
    return get_cotacao_comparacao(False, FUNDOS_TICKER)


def get_cotacao_comparacao(alta: bool, tickers: []) -> CotacaoAltaOuBaixaSchema:
    ativo = yf.download(tickers, period="1d")
    variacao = ((ativo['Close'] - ativo['Open']) / ativo['Open']) * 100

    list_ativo = []
    for k in variacao.keys():
        for item in variacao.get(key=k):
            if not math.isnan(item) and not math.isinf(item):
                list_ativo.append(CotacaoAltaOuBaixaSchema(codigo=k, valor=item))

    list_ativo.sort(key=lambda x: x.valor, reverse=alta)
    return list_ativo[0:20]


def get_preco_abertura_acoes():
    workbook = load_workbook("src/analise/rendavariavel.xlsx")
    sheet = workbook["Ações"]
    row_count = sheet.max_row
    tickers = []
    for row in range(2, row_count + 1):
        ticker = sheet.cell(row=row, column=1).value
        tickers.append(f'{ticker}.SA')

    dowload_tickers = yf.download(tickers, period='1d')['Close']

    for ticker in tickers:
        print(dowload_tickers[ticker][0])
