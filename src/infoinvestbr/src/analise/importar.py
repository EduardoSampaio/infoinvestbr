import models
from models import FundosImobiliario, Acao
from openpyxl import load_workbook

from sqlalchemy.orm import Session


def import_fundos_imobiliarios(db: Session):
    workbook = load_workbook("src/analise/rendavariavel.xlsx")
    sheet = workbook["FIIS"]
    row_count = sheet.max_row
    list_fii = []
    for row in range(2, row_count + 1):
        fii = FundosImobiliario(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=5).value,
            sheet.cell(row=row, column=6).value,
            sheet.cell(row=row, column=7).value,
            sheet.cell(row=row, column=8).value,
            sheet.cell(row=row, column=9).value,
            sheet.cell(row=row, column=10).value,
            sheet.cell(row=row, column=11).value,
            sheet.cell(row=row, column=12).value,
            sheet.cell(row=row, column=13).value,
            sheet.cell(row=row, column=14).value,
            sheet.cell(row=row, column=15).value,
            sheet.cell(row=row, column=16).value,
            sheet.cell(row=row, column=17).value,
            sheet.cell(row=row, column=18).value,
            sheet.cell(row=row, column=19).value,
        )
        list_fii.append(fii)

    count = db.query(models.FundosImobiliario).count()
    if count == 0:
        db.add_all(list_fii)
        db.commit()


def import_acoes(db: Session):
    workbook = load_workbook("src/analise/rendavariavel.xlsx")
    sheet = workbook["Ações"]
    row_count = sheet.max_row
    list_acoes = []
    for row in range(2, row_count + 1):
        acoes = Acao(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=5).value,
            sheet.cell(row=row, column=6).value,
            sheet.cell(row=row, column=7).value,
            sheet.cell(row=row, column=8).value,
            sheet.cell(row=row, column=9).value,
            sheet.cell(row=row, column=10).value,
            sheet.cell(row=row, column=11).value,
            sheet.cell(row=row, column=12).value,
            sheet.cell(row=row, column=13).value,
            sheet.cell(row=row, column=14).value,
            sheet.cell(row=row, column=15).value,
            sheet.cell(row=row, column=16).value,
            sheet.cell(row=row, column=17).value,
            sheet.cell(row=row, column=18).value,
            sheet.cell(row=row, column=19).value,
            sheet.cell(row=row, column=20).value,
            sheet.cell(row=row, column=21).value
        )
        list_acoes.append(acoes)

    count = db.query(models.Acao).count()
    if count == 0:
        db.add_all(list_acoes)
        db.commit()


def read_codigos_file():
    workbook = load_workbook("src/analise/rendavariavel.xlsx")
    sheet_acoes = workbook["Ações"]
    row_count_acoes = sheet_acoes.max_row
    sheet_fundos = workbook["FIIS"]
    row_count_fundos = sheet_fundos.max_row

    list_codigos = []
    for row in range(2, row_count_acoes + 1):
        codigo = sheet_acoes.cell(row=row, column=1).value
        list_codigos.append(codigo)

    for row in range(2, row_count_fundos + 1):
        codigo = sheet_fundos.cell(row=row, column=1).value
        list_codigos.append(codigo)

    return list_codigos
