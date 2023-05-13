from openpyxl import load_workbook
from sqlalchemy.orm import Session
from src.analise.models import Acao, FundosImobiliario
from src.analise.schemas import AcaoRequestSchema, AcaoResponseSchema, FundosImobiliarioRequestSchema, \
    FundosImobiliarioResponseSchema
import logging as logger


from src.core.database import SessionLocal
from src.core.models import CotacaoAtivo


def convert_to_schema_acao(model: Acao):
    if model is None:
        return {}

    db = SessionLocal()
    cotacao = db.query(CotacaoAtivo.preco).filter(CotacaoAtivo.codigo == model.codigo).first()
    db.close()

    response = AcaoResponseSchema()
    response.id = model.id
    response.codigo = model.codigo
    response.pl = formatar_numeros(model.pl)
    response.pvp = formatar_numeros(model.pvp, 4)
    response.psr = formatar_numeros(model.psr, 4)
    response.dividend_yield = formatar_numeros(model.dividend_yield, porcentagem=True)
    response.p_ativo = formatar_numeros(model.p_ativo)
    response.p_cap_giro = formatar_numeros(model.p_cap_giro)
    response.p_ebit = formatar_numeros(model.p_ebit)
    response.p_ativ_circ_liq = formatar_numeros(model.p_ativ_circ_liq)
    response.ev_ebit = formatar_numeros(model.ev_ebit)
    response.ev_ebitda = formatar_numeros(model.ev_ebitda)
    response.margem_ebit = formatar_numeros(model.margem_ebit, porcentagem=True)
    response.margem_liquida = formatar_numeros(model.margem_liquida, porcentagem=True)
    response.liq_corrente = formatar_numeros(model.liq_corrente)
    response.roic = formatar_numeros(model.roic, porcentagem=True)
    response.roe = formatar_numeros(model.roe, porcentagem=True)
    response.liq_2meses = formatar_numeros(model.liq_2meses)
    response.patrimonio_liquido = formatar_numeros(model.patrimonio_liquido)
    response.div_bruta_patrim = formatar_numeros(model.div_bruta_patrim)
    response.cresc_rec_5a = formatar_numeros(model.cresc_rec_5a, porcentagem=True)
    response.preco = formatar_numeros(cotacao.preco, 2)
    response.lpa = model.lpa
    response.vpa = model.vpa
    response.setor = model.setor
    response.tipo = model.tipo
    response.nome = model.nome
    response.imagem = model.imagem
    response.descricao = model.descricao
    response.cnpj = model.cnpj
    response.sub_setor = model.sub_setor

    return response


def convert_to_schema_fundos(model: FundosImobiliario):
    if model is None:
        return {}

    db = SessionLocal()
    cotacao = db.query(CotacaoAtivo).filter(CotacaoAtivo.codigo == model.codigo).first()
    db.close()
    response = FundosImobiliarioResponseSchema()
    response.id = model.id
    response.codigo = model.codigo
    response.nome = model.nome
    response.descricao = model.descricao
    response.administrador = model.administrador
    response.cnpj = model.cnpj
    response.setor = model.setor
    response.liquidez_diaria = model.liquidez_diaria
    response.quantidade_ativos = model.quantidade_ativos
    response.taxa_administracao = model.taxa_administracao
    response.taxa_gestao = model.taxa_gestao
    response.taxa_performance = model.taxa_performance
    response.tipo_gestao = model.tipo_gestao
    response.dividendo = formatar_numeros(model.dividendo, 4)
    response.dividend_yield = formatar_numeros(model.dividend_yield, 4, True)
    response.dy_ano = formatar_numeros(model.dy_ano, 4, True)
    response.variacao_preco = formatar_numeros(model.variacao_preco, 4, True)
    response.rentab_periodo = formatar_numeros(model.rentab_periodo, 4, True)
    response.rentab_acumulada = formatar_numeros(model.rentab_acumulada, 4, True)
    response.patrimonio_liq = formatar_numeros(model.patrimonio_liq, 4)
    response.vpa = formatar_numeros(model.vpa, 4)
    response.p_vpa = formatar_numeros(model.p_vpa, 4)
    response.dy_patrimonial = formatar_numeros(model.dy_patrimonial, 4, True)
    response.variacao_patrimonial = formatar_numeros(model.variacao_patrimonial, 4, True)
    response.rentab_patr_no_periodo = formatar_numeros(model.rentab_patr_no_periodo, 4, True)
    response.rentab_patr_acumulada = formatar_numeros(model.rentab_patr_acumulada, 4, True)
    response.vacancia_fisica = formatar_numeros(model.vacancia_fisica, 4, True)
    response.vacancia_financeira = formatar_numeros(model.vacancia_financeira, 4, True)
    response.preco = formatar_numeros(cotacao.preco, 2)

    return response


def get_acoes(db: Session, skip: int = 0, limit: int = 100):
    acoes = db.query(Acao).order_by(Acao.codigo).offset(skip).limit(limit).all()
    return acoes


def get_acoes_by_setor(db: Session, setor: str):
    acoes = db.query(Acao).filter(Acao.setor == setor).all()
    return acoes


def get_fundos_imobiliarios(db: Session, skip: int = 0, limit: int = 100):
    fundos = db.query(FundosImobiliario).order_by(FundosImobiliario.codigo).offset(skip).limit(limit).all()
    return fundos


def get_fundos_imobiliarios_setor(db: Session, setor: str):
    fundos = db.query(FundosImobiliario).filter(FundosImobiliario.setor == setor).all()
    return fundos


def get_acoes_by_id(db: Session, acao_id: int):
    return db.query(Acao).filter(Acao.id == acao_id).first()


def get_fundos_imobiliarios_by_id(db: Session, fundos_id: int):
    return db.query(FundosImobiliario).filter(FundosImobiliario.id == fundos_id).first()


def get_acoes_by_codigo(db: Session, codigo: str):
    return convert_to_schema_acao(db.query(Acao).filter(Acao.codigo == codigo).first())


def get_fundos_imobiliarios_by_codigo(db: Session, codigo: str):
    fundo = db.query(FundosImobiliario).filter(FundosImobiliario.codigo == codigo).first()
    return convert_to_schema_fundos(fundo)


def update_acoes(db: Session, acao: AcaoRequestSchema):
    _acao = get_acoes_by_id(db, acao.id)
    _acao.codigo = acao.codigo
    _acao.tipo = acao.tipo
    _acao.nome = acao.nome
    _acao.imagem = acao.imagem
    _acao.lpa = acao.lpa
    _acao.vpa = acao.vpa
    _acao.setor = acao.setor
    _acao.descricao = acao.descricao
    _acao.cresc_rec_5a = acao.cresc_rec_5a
    _acao.div_bruta_patrim = acao.div_bruta_patrim
    _acao.ev_ebit = acao.ev_ebit
    _acao.dividend_yield = acao.dividend_yield
    _acao.liq_2meses = acao.liq_2meses
    _acao.pl = acao.pl
    _acao.pvp = acao.pvp
    _acao.psr = acao.psr
    _acao.p_ativo = acao.p_ativo
    _acao.p_cap_giro = _acao.p_cap_giro
    _acao.p_ebit = _acao.p_ebit
    _acao.p_ativ_circ_liq = acao.p_ativ_circ_liq
    _acao.ev_ebitda = acao.ev_ebitda
    _acao.margem_ebit = acao.margem_ebit
    _acao.margem_liquida = acao.margem_liquida
    _acao.liq_corrente = acao.liq_corrente
    _acao.roic = acao.roic
    _acao.roe = acao.roe
    _acao.patrimonio_liquido = acao.patrimonio_liquido

    db.commit()
    db.refresh(_acao)


def update_fundos(db: Session, fundo: FundosImobiliarioRequestSchema):
    _fundo = get_fundos_imobiliarios_by_id(db, fundo.id)
    _fundo.nome = fundo.nome
    _fundo.descricao = fundo.descricao
    _fundo.administrador = fundo.administrador
    _fundo.cnpj = fundo.cnpj
    _fundo.taxa_administracao = fundo.taxa_administracao
    _fundo.taxa_gestao = fundo.taxa_gestao
    _fundo.taxa_performance = fundo.taxa_performance
    _fundo.tipo_gestao = fundo.tipo_gestao
    _fundo.setor = fundo.setor
    _fundo.liquidez_diaria = fundo.liquidez_diaria
    _fundo.dividendo = fundo.dividendo
    _fundo.dividend_yield = fundo.dividend_yield
    _fundo.dy_ano = fundo.dy_ano
    _fundo.variacao_preco = fundo.variacao_preco
    _fundo.rentab_periodo = fundo.rentab_periodo
    _fundo.rentab_acumulada = fundo.rentab_acumulada
    _fundo.patrimonio_liq = fundo.patrimonio_liq
    _fundo.vpa = fundo.vpa
    _fundo.p_vpa = fundo.p_vpa
    _fundo.dy_patrimonial = fundo.dy_patrimonial
    _fundo.variacao_patrimonial = fundo.variacao_patrimonial
    _fundo.rentab_patr_no_periodo = fundo.rentab_patr_no_periodo
    _fundo.rentab_patr_acumulada = fundo.rentab_patr_acumulada
    _fundo.vacancia_fisica = fundo.vacancia_fisica
    _fundo.vacancia_financeira = fundo.vacancia_financeira
    _fundo.quantidade_ativos = fundo.quantidade_ativos

    db.commit()
    db.refresh(_fundo)


def import_fundos_imobiliarios(db: Session):
    logger.info("Inicio da Importação de fundos Imobiliários")
    workbook = load_workbook("src/analise/rendavariavel.xlsx")
    sheet = workbook["FIIS"]
    row_count = sheet.max_row
    list_fii = []
    for row in range(2, row_count + 1):
        fii = FundosImobiliario(
            codigo_do_fundo=sheet.cell(row=row, column=1).value,
            setor=sheet.cell(row=row, column=2).value,
            liquidez_diaria=sheet.cell(row=row, column=3).value,
            dividendo=sheet.cell(row=row, column=4).value,
            dividend_yield=sheet.cell(row=row, column=5).value,
            dy_ano=sheet.cell(row=row, column=6).value,
            variacao_preco=sheet.cell(row=row, column=7).value,
            rentab_periodo=sheet.cell(row=row, column=8).value,
            rentab_acumulada=sheet.cell(row=row, column=9).value,
            patrimonio_liq=sheet.cell(row=row, column=10).value,
            vpa=sheet.cell(row=row, column=11).value,
            p_vpa=sheet.cell(row=row, column=12).value,
            dy_patrimonial=sheet.cell(row=row, column=13).value,
            variacao_patrimonial=sheet.cell(row=row, column=14).value,
            rentab_patr_no_periodo=sheet.cell(row=row, column=15).value,
            rentab_patr_acumulada=sheet.cell(row=row, column=16).value,
            vacancia_fisica=sheet.cell(row=row, column=17).value,
            vacancia_financeira=sheet.cell(row=row, column=18).value,
            quantidade_ativos=sheet.cell(row=row, column=19).value,
            nome=sheet.cell(row=row, column=20).value,
            cnpj=sheet.cell(row=row, column=21).value,
            administrador=sheet.cell(row=row, column=22).value,
        )
        list_fii.append(fii)

    count = db.query(FundosImobiliario).count()
    if count == 0:
        db.add_all(list_fii)
        db.commit()
    logger.info("Fim da Importação de fundos Imobiliários")


def import_acoes(db: Session):
    workbook = load_workbook("src/analise/rendavariavel.xlsx")
    sheet = workbook["Ações"]
    row_count = sheet.max_row
    list_acoes = []
    for row in range(2, row_count + 1):
        acoes = Acao(
            codigo=sheet.cell(row=row, column=1).value,
            pl=sheet.cell(row=row, column=2).value,
            pvp=sheet.cell(row=row, column=3).value,
            psr=sheet.cell(row=row, column=4).value,
            dividend_yield=sheet.cell(row=row, column=5).value,
            p_ativo=sheet.cell(row=row, column=6).value,
            p_cap_giro=sheet.cell(row=row, column=7).value,
            p_ebit=sheet.cell(row=row, column=8).value,
            p_ativ_circ_liq=sheet.cell(row=row, column=9).value,
            ev_ebit=sheet.cell(row=row, column=10).value,
            ev_ebitda=sheet.cell(row=row, column=11).value,
            margem_ebit=sheet.cell(row=row, column=12).value,
            margem_liquida=sheet.cell(row=row, column=13).value,
            liq_corrente=sheet.cell(row=row, column=14).value,
            roic=sheet.cell(row=row, column=15).value,
            roe=sheet.cell(row=row, column=16).value,
            liq_2meses=sheet.cell(row=row, column=17).value,
            patrimonio_liquido=sheet.cell(row=row, column=18).value,
            div_bruta_patrim=sheet.cell(row=row, column=19).value,
            cresc_rec_5a=sheet.cell(row=row, column=20).value,
            setor=sheet.cell(row=row, column=21).value,
            sub_setor=sheet.cell(row=row, column=22).value,
            nome=sheet.cell(row=row, column=23).value,
            cnpj=sheet.cell(row=row, column=24).value,
            imagem=sheet.cell(row=row, column=25).value
        )
        list_acoes.append(acoes)

    count = db.query(Acao).count()
    if count == 0:
        db.add_all(list_acoes)
        db.commit()


def remove_todas_acoes(db: Session):
    acoes = db.query(Acao).all()

    for acao in acoes:
        db.delete(acao)

    db.commit()


def remove_todos_fundos(db: Session):
    fundos = db.query(FundosImobiliario).all()

    for fundo in fundos:
        db.delete(fundo)

    db.commit()


def formatar_numeros(valor: float, decimais: int = 4, porcentagem: bool = False) -> float:
    if valor is None:
        return 0.0
    valor_formatado = round(valor, decimais)
    if porcentagem:
        return valor_formatado * 100
    else:
        return valor_formatado
